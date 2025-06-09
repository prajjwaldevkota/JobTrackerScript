# === HELPERS ===
import imaplib
import email
import re
import os
import openpyxl
from email.header import decode_header
from datetime import datetime
from email.utils import parsedate_to_datetime
from dotenv import load_dotenv

# Configuration constants
load_dotenv()
EMAIL = os.getenv("EMAIL")
APP_PASSWORD = os.getenv("APP_PASSWORD")
IMAP_SERVER = "imap.gmail.com"
EXCEL_FILE = "job_applications_cleaned.xlsx"
DATE_LIMIT = "28-May-2025"


def connect_email():
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL, APP_PASSWORD)
    return mail


def parse_subject(subj):
    try:
        decoded, enc = decode_header(subj)[0]
        return decoded.decode(enc or 'utf-8') if isinstance(decoded, bytes) else decoded
    except:
        return subj or ""


def get_email_body(msg):
    """Extract the email body content from the message"""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition"))

            # Skip attachments
            if "attachment" in content_disposition:
                continue

            if content_type == "text/plain" or content_type == "text/html":
                try:
                    body = part.get_payload(decode=True).decode()
                    return body
                except:
                    pass
    else:
        # Not multipart - get payload directly
        try:
            body = msg.get_payload(decode=True).decode()
            return body
        except:
            pass

    return ""


def extract_job_info(subject, sender, body=None):
    sl = subject.lower()
    body_text = body.lower() if body else ""

    # Rejection phrases in email body
    rejection_body_phrases = [
        "at this time we are not proceeding",
        "we are not proceeding with your application",
        "unfortunately",
        "we regret to inform you",
        "we have decided to move forward with other candidates",
        "we will not be progressing",
        "not selected",
        "not successful",
        "we've decided not to move forward",
        "thank you for your interest",
        "position has been filled",
        "no longer being considered",
        "we have chosen another candidate",
        "not a match",
        "not the right fit"
    ]

    # give status a default
    status = "Applied"

    # Subject-based status detection
    if body and any(phrase in body_text for phrase in rejection_body_phrases):
        status = "Rejected"
        print(f"Found rejection phrase in email body for: {subject}")
    elif any(k in sl for k in ["rejected", "unsuccessful", "not selected", "not moving forward", "not proceeding", "thank you for the interest"]):
        status = "Rejected"
    elif any(k in sl for k in ["offer", "congratulations", "welcome aboard", "pleased to offer"]):
        status = "Offer"
    elif any(k in sl for k in ["interview", "shortlisted", "next round", "next steps", "assessment", "interview scheduled", "interview confirmation"]):
        status = "Interview"

    # Job title extraction
    title = subject  # Default to full subject

    # Try to identify the company name first
    company_patterns = [
        r'at\s+["\']?([\w\s&\-\.]+?)["\']?(?:$|:|\.|\s\()',  # "at [company]"
        # "from [company]"
        r'from\s+["\']?([\w\s&\-\.]+?)["\']?(?:$|:|\.|\s\()',
    ]

    comp = None
    for pattern in company_patterns:
        cm = re.search(pattern, subject, re.IGNORECASE)
        if cm:
            comp = cm.group(1).strip()
            break

    # If company was found, try to extract title by removing common phrases and the company
    if comp:
        # Common prefixes/suffixes to clean up
        to_remove = [
            f"at {comp}",
            f"from {comp}",
            "application for",
            "thank you for your application",
            "application received",
            "your application for",
            "re:",
            "thank you for applying",
            "application confirmation",
            "application was sent",
        ]

        cleaned_title = subject
        for phrase in to_remove:
            cleaned_title = re.sub(re.escape(phrase), "",
                                   cleaned_title, flags=re.IGNORECASE)

        # Clean up extra whitespace and punctuation
        title = re.sub(r'^\W+|\W+$', '', cleaned_title).strip()

        # If title became empty, revert to subject
        if not title:
            title = subject

    # Fallback to email domain if no company found
    if not comp:
        try:
            email_parts = sender.split('<')[-1].split('>')[0].split('@')
            if len(email_parts) > 1:
                domain = email_parts[1]
                comp = domain.split('.')[0].capitalize()
        except:
            comp = "Unknown"

    return title.strip(), comp.strip(), status


def should_update_status(current_status, new_status):
    """Determine if status should be updated based on hierarchy of statuses"""
    status_hierarchy = {
        "Applied": 0,
        "Interview": 1,
        "Offer": 2,
        "Rejected": 3  # Rejection is final state
    }

    # Only update if new status is "higher" in the hierarchy
    return status_hierarchy.get(new_status, 0) > status_hierarchy.get(current_status, 0)


# === MAIN ===
def fetch_jobs():
    # Load existing workbook and prepare data structures
    companies_map = {}  # Company name -> list of row indices
    existing_uids = set()
    uid_column_index = 5  # 0-based index for UID column (column F)

    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        headers = [c.value for c in sheet[1]]

        # Ensure UID column
        if "UID" not in headers:
            sheet.cell(row=1, column=len(headers)+1, value="UID")
            wb.save(EXCEL_FILE)

        # Load existing data and build lookup maps
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Handle variable number of columns
            row_data = list(row) + [None] * (6 - len(row))
            date_val, title_val, comp_val, status_val, sender_val, uid_val = row_data

            if uid_val:
                existing_uids.add(str(uid_val))

            # Build company name -> row index map for quick lookups
            if comp_val:
                comp_key = comp_val.strip().lower()
                if comp_key not in companies_map:
                    companies_map[comp_key] = []
                companies_map[comp_key].append(row_idx)
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Date", "Job Title", "Company",
                     "Status", "Sender", "UID"])
        wb.save(EXCEL_FILE)

    try:
        # Connect and search emails
        mail = connect_email()
        mail.select("INBOX")

        # Try category:primary first
        typ, data = mail.uid(
            'search', None, f'SINCE {DATE_LIMIT} X-GM-RAW "Category:Primary"')

        # If it fails or finds nothing, fall back to inbox search
        if typ != 'OK' or not data[0]:
            print("⚠️ Falling back to basic INBOX search.")
            typ, data = mail.uid('search', None, f'SINCE "{DATE_LIMIT}"')

        uids = data[0].split()
        print(f"🔍 Found {len(uids)} messages in Primary since {DATE_LIMIT}")

        # Simple inclusion keywords to filter relevant emails
        include_kw = [
            "thank you for your application",
            "application received",
            "your application for",
            "we've received your application",
            "interview",
            "shortlisted",
            "internship",
            "technical support",
            "support analyst",
            "application was sent",
            "Your online application has been successfully submitted",
            "Help Desk Analyst", "Junior Software",
            "Thank You for Applying",
        ]

        # Blacklist - more specific patterns to exclude
        blacklist = [
            "visa", "ircc", "immigration",
            "30+ new tech internships posted this week",
            "job alert", "apply now",
            "sql interview challenge",
            "is hiring", "new internships",
            "internship", "hiring",
            "job opportunity", "one of the first",
            "be a great fit", "webinar",
            "newsletter", "unsubscribe",
            "career fair", "notification", "event", "first", "apply to", "actively recruiting", "you would be a great fit", "and more",
            "glassdoor", "juli"
        ]

        new_rows = []
        updated_rows = []  # Track which rows were updated

        for uid in uids:
            uid_str = uid.decode() if isinstance(uid, bytes) else str(uid)
            if uid_str in existing_uids:
                continue

            try:
                typ, msg_data = mail.uid('fetch', uid, '(RFC822)')
                if typ != 'OK' or not msg_data or not msg_data[0]:
                    continue

                msg = email.message_from_bytes(msg_data[0][1])

                subj = parse_subject(msg.get('Subject', ''))
                sl = subj.lower()
                sender = msg.get('From', '')

                # Extract email body for additional rejection analysis
                body = get_email_body(msg)

                # Check inclusion criteria - must have one of the keywords
                if not any(kw.lower() in sl for kw in include_kw):
                    continue

                # Improved blacklist filtering - check each word individually
                # Skip if ANY blacklist term matches
                blacklisted = False
                for term in blacklist:
                    if term.lower() in sl:
                        print(
                            f"⏭️ Skipping blacklisted: {subj} [matched: {term}]")
                        blacklisted = True
                        break

                if blacklisted:
                    continue

                # Parse actual sent date
                try:
                    dt = parsedate_to_datetime(msg.get('Date', ''))
                    date_str = dt.strftime('%Y-%m-%d')
                except:
                    date_str = datetime.now().strftime('%Y-%m-%d')

                title, comp, status = extract_job_info(subj, sender, body)
                comp_key = comp.lower()

                # Check if we already have an application for this company
                updated = False
                if comp_key in companies_map:
                    # Look through all entries for this company
                    for row_idx in companies_map[comp_key]:
                        # Get current data from spreadsheet
                        current_data = [sheet.cell(
                            row=row_idx, column=i).value for i in range(1, 7)]
                        current_date, current_title, current_company, current_status, current_sender, current_uid = current_data

                        # If status is "Applied" or "Interview" and new status is more definitive, update it
                        if current_status and should_update_status(current_status, status):
                            # Update status and add new UID
                            sheet.cell(row=row_idx, column=4,
                                       value=status)  # Update status
                            sheet.cell(row=row_idx, column=6,
                                       value=uid_str)  # Update UID

                            print(
                                f"🔄 Updated: {current_title} at {current_company} from '{current_status}' to '{status}'")
                            updated_rows.append(row_idx)
                            existing_uids.add(uid_str)
                            updated = True
                            break

                # If not updated, add as a new entry
                if not updated:
                    new_rows.append(
                        (date_str, title, comp, status, sender, uid_str))
                    existing_uids.add(uid_str)
                    print(f"📌 New job: {title} at {comp} ({status}) - {subj}")

            except Exception as e:
                print(f"⚠️ Error processing email UID {uid_str}: {str(e)}")

        mail.logout()

        # Save changes to workbook
        if new_rows or updated_rows:
            # Add new rows
            for row in new_rows:
                sheet.append(row)

            wb.save(EXCEL_FILE)
            print(
                f"✅ Appended {len(new_rows)} new jobs and updated {len(updated_rows)} existing entries.")
        else:
            print("ℹ️ No new jobs to add or update.")

    except Exception as e:
        print(f"❌ Error: {str(e)}")


if __name__ == '__main__':
    fetch_jobs()
